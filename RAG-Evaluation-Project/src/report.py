"""Final comparison dashboard (terminal), Excel export, and JSON metrics export
(for regression testing) across all RAG patterns."""
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src import config


def print_comparison(all_results):
    """Print a side-by-side comparison table of all evaluated patterns."""
    print("\n" + "=" * 100)
    print("  RAG PATTERN COMPARISON — FINAL RESULTS".center(100))
    print("=" * 100)

    # Retrieval metrics table
    retrieval_keys = list(all_results[0]["retrieval_metrics"].keys())
    header = f"  {'Pattern':<42s}" + "".join(f"{k:>12s}" for k in retrieval_keys)
    print("\nRETRIEVAL METRICS")
    print(header)
    print("-" * len(header))
    for r in all_results:
        row = f"  {r['pattern']:<42s}"
        row += "".join(f"{r['retrieval_metrics'][k]:>12.4f}" for k in retrieval_keys)
        print(row)

    # Generation metrics table
    gen_keys = list(all_results[0]["generation_metrics"].keys())
    header = f"  {'Pattern':<42s}" + "".join(f"{k:>22s}" for k in gen_keys)
    print(f"\nGENERATION METRICS (LLM-as-judge, sample size = {all_results[0]['sample_size']})")
    print(header)
    print("-" * len(header))
    for r in all_results:
        row = f"  {r['pattern']:<42s}"
        row += "".join(f"{r['generation_metrics'][k]:>22.4f}" for k in gen_keys)
        print(row)

    # RAGAS metrics table (if available)
    if all(r.get("ragas_metrics") for r in all_results):
        ragas_keys = list(all_results[0]["ragas_metrics"].keys())
        header = f"  {'Pattern':<42s}" + "".join(f"{k:>20s}" for k in ragas_keys)
        print("\nRAGAS METRICS")
        print(header)
        print("-" * len(header))
        for r in all_results:
            row = f"  {r['pattern']:<42s}"
            row += "".join(f"{r['ragas_metrics'][k]:>20.4f}" for k in ragas_keys)
            print(row)

    # Timing
    print("\nTIMING")
    print(f"  {'Pattern':<42s} {'Build (s)':>12s} {'Total (s)':>12s}")
    for r in all_results:
        print(f"  {r['pattern']:<42s} {r['build_time_sec']:>12.1f} {r['total_time_sec']:>12.1f}")

    # Key differences summary
    print("\nKEY DIFFERENCES")
    best_hit = max(all_results, key=lambda r: r["retrieval_metrics"].get(f"Hit Rate@{config.TOP_K}", 0))
    best_grounded = max(all_results, key=lambda r: r["generation_metrics"].get("Groundedness (1-5)", 0))
    least_hallucination = min(all_results, key=lambda r: r["generation_metrics"].get("Hallucination Rate (0-1)", 1))
    fastest = min(all_results, key=lambda r: r["total_time_sec"])
    print(f"  Best retrieval (Hit Rate@{config.TOP_K}):  {best_hit['pattern']}")
    print(f"  Most grounded answers:        {best_grounded['pattern']}")
    print(f"  Lowest hallucination rate:    {least_hallucination['pattern']}")
    print(f"  Fastest end-to-end run:       {fastest['pattern']}")

    if config.TRACING_ENABLED or config.LANGFUSE_TRACING_ENABLED:
        print("\nOBSERVABILITY")
        if config.TRACING_ENABLED:
            print(f"  LangChain-based pattern runs (1-5, 7) traced to LangSmith project '{config.LANGSMITH_PROJECT}'.")
            print(f"    View traces: {config.langsmith_project_url()}")
        if config.LANGFUSE_TRACING_ENABLED:
            print("  All 7 patterns (including the raw OpenAI SDK File Search pattern) traced to Langfuse.")
            print(f"    View traces: {config.langfuse_project_url()}")

    print("=" * 100)


def export_to_excel(all_results, output_path=None):
    """Export the full comparison to a multi-sheet Excel workbook."""
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(config.OUTPUT_DIR, f"{config.EXPERIMENT_NAME}_{timestamp}.xlsx")

    wb = Workbook()
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    def write_table(ws, title, columns, rows):
        ws["A1"] = title
        ws["A1"].font = Font(name="Arial", bold=True, size=13)
        for c, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=3, column=c, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r, row_values in enumerate(rows, start=4):
            for c, value in enumerate(row_values, start=1):
                ws.cell(row=r, column=c, value=value)
        for c in range(1, len(columns) + 1):
            ws.column_dimensions[chr(64 + c)].width = 28

    # Sheet 1: Summary (config)
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"RAG Pattern Comparison — {config.EXPERIMENT_NAME}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws["A2"] = f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    config_rows = [
        ("Embedding Model", config.EMBEDDING_MODEL),
        ("Generator LLM", config.GENERATOR_LLM),
        ("Judge LLM", config.JUDGE_LLM),
        ("Chunk Size / Overlap", f"{config.CHUNK_SIZE} / {config.CHUNK_OVERLAP}"),
        ("Top K", config.TOP_K),
        ("E2E Sample Size", config.E2E_SAMPLE_SIZE),
        ("Dataset Size", all_results[0]["dataset_size"]),
    ]
    write_table(ws, "Configuration", ["Setting", "Value"], config_rows)

    # Sheet 2: Retrieval metrics
    ws = wb.create_sheet("Retrieval Metrics")
    retrieval_keys = list(all_results[0]["retrieval_metrics"].keys())
    rows = [[r["pattern"]] + [r["retrieval_metrics"][k] for k in retrieval_keys] for r in all_results]
    write_table(ws, "Retrieval Metrics", ["Pattern"] + retrieval_keys, rows)

    # Sheet 3: Generation metrics
    ws = wb.create_sheet("Generation Metrics")
    gen_keys = list(all_results[0]["generation_metrics"].keys())
    rows = [[r["pattern"]] + [r["generation_metrics"][k] for k in gen_keys] for r in all_results]
    write_table(ws, "Generation Metrics (LLM-as-judge)", ["Pattern"] + gen_keys, rows)

    # Sheet 4: RAGAS metrics (if available)
    if all(r.get("ragas_metrics") for r in all_results):
        ws = wb.create_sheet("RAGAS Metrics")
        ragas_keys = list(all_results[0]["ragas_metrics"].keys())
        rows = [[r["pattern"]] + [r["ragas_metrics"][k] for k in ragas_keys] for r in all_results]
        write_table(ws, "RAGAS Metrics", ["Pattern"] + ragas_keys, rows)

    # Sheet 5: Timing
    ws = wb.create_sheet("Timing")
    rows = [[r["pattern"], r["build_time_sec"], r["total_time_sec"]] for r in all_results]
    write_table(ws, "Timing", ["Pattern", "Build Time (s)", "Total Time (s)"], rows)

    # Sheet 6: Per-query generation scores
    ws = wb.create_sheet("Per-Query Generation")
    rows = []
    for r in all_results:
        for q in r["per_query_generation"]:
            rows.append(
                [
                    r["pattern"],
                    q["question_id"],
                    q["answer_relevance"],
                    q["groundedness"],
                    q["hallucination_rate"],
                    q["coherence"],
                ]
            )
    write_table(
        ws,
        "Per-Query Generation Scores",
        ["Pattern", "Question ID", "Answer Relevance", "Groundedness", "Hallucination Rate", "Coherence"],
        rows,
    )

    wb.save(output_path)
    print(f"\nFull results exported to: {output_path}")
    return output_path


def export_to_json(all_results, output_path=None):
    """Export per-pattern retrieval/generation/RAGAS metrics + timing to JSON.

    Used as the input to `scripts/regression_check.py`, which compares this
    against a saved baseline (`outputs/baseline.json`) to catch quality
    regressions (e.g. a drop in Hit Rate or a rise in Hallucination Rate).
    """
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    if output_path is None:
        output_path = os.path.join(config.OUTPUT_DIR, "latest_results.json")

    summary = {
        "timestamp": datetime.now().isoformat(),
        "dataset_size": all_results[0]["dataset_size"],
        "sample_size": all_results[0]["sample_size"],
        "patterns": {
            r["pattern"]: {
                "retrieval_metrics": r["retrieval_metrics"],
                "generation_metrics": r["generation_metrics"],
                "ragas_metrics": r.get("ragas_metrics"),
                "build_time_sec": r["build_time_sec"],
                "total_time_sec": r["total_time_sec"],
            }
            for r in all_results
        },
    }

    with open(output_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"Metrics snapshot exported to: {output_path}")
    return output_path
